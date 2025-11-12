import cv2
import numpy as np
from ultralytics import YOLO
import torch
import torchvision.transforms as T
from torchvision.models import mobilenet_v2, MobileNet_V2_Weights
from collections import defaultdict
import time
import requests
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PILImage
import json
 
# ================================
# CONFIG
# ================================
URLS_FILE = "Image_url.txt"
CONFIDENCE_THRESH = 0.45
 
# ================================
# 1. MODELS & SETUP
# ================================
 
# YOLO for car detection
car_model = YOLO('yolo11n.pt')
 
# Deep learning orientation classifier (front/rear)
device = 'cuda' if torch.cuda.is_available() else 'cpu'
orientation_model = mobilenet_v2(weights=MobileNet_V2_Weights.IMAGENET1K_V1)
orientation_model.classifier[1] = torch.nn.Linear(orientation_model.classifier[1].in_features, 2)
# 🔔 TODO: Load fine-tuned weights if available:
# orientation_model.load_state_dict(torch.load('car_orientation.pth', map_location=device))
orientation_model = orientation_model.to(device).eval()
 
transform = T.Compose([
    T.ToPILImage(),
    T.Resize((224, 224)),
    T.ToTensor(),
    T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
])
 
# Motion tracking (simulated per image)
car_tracks = defaultdict(list)
next_id = 0
 
# ================================
# 2. HELPER FUNCTIONS (same as before)
# ================================
 
def get_orientation_dl(car_crop):
    try:
        tensor = transform(car_crop).unsqueeze(0).to(device)
        with torch.no_grad():
            out = orientation_model(tensor)
            pred = torch.argmax(out, dim=1).item()
        return "rear" if pred == 1 else "front"
    except:
        return "unknown"
 
def get_orientation_features(car_crop):
    h, w = car_crop.shape[:2]
    if h == 0 or w == 0:
        return "unknown"
    light_region = car_crop[int(0.6 * h):, :]
    if light_region.size == 0:
        return "unknown"
 
    hsv = cv2.cvtColor(light_region, cv2.COLOR_BGR2HSV)
    gray = cv2.cvtColor(light_region, cv2.COLOR_BGR2GRAY)
 
    red1 = cv2.inRange(hsv, (0, 100, 100), (10, 255, 255))
    red2 = cv2.inRange(hsv, (170, 100, 100), (180, 255, 255))
    red_pixels = cv2.countNonZero(red1 + red2)
 
    _, bright = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY)
    bright_pixels = cv2.countNonZero(bright)
 
    if red_pixels > 60:
        return "rear"
    elif bright_pixels > 120:
        return "front"
    else:
        return "unknown"
 
def assign_track_id(box, max_dist=50):
    global next_id
    cx = (box[0] + box[2]) / 2
    cy = (box[1] + box[3]) / 2
    centroid = (cx, cy)
 
    best_id = None
    min_dist = float('inf')
    for tid, track in car_tracks.items():
        if track:
            last_cx, last_cy = track[-1]
            dist = np.sqrt((cx - last_cx)**2 + (cy - last_cy)**2)
            if dist < min_dist and dist < max_dist:
                min_dist = dist
                best_id = tid
 
    if best_id is None:
        best_id = next_id
        next_id += 1
 
    car_tracks[best_id].append(centroid)
    if len(car_tracks[best_id]) > 5:
        car_tracks[best_id].pop(0)
 
    return best_id
 
def get_orientation_motion(track_id):
    track = car_tracks.get(track_id, [])
    if len(track) < 2:
        return "unknown"
 
    start_x = track[0][0]
    end_x = track[-1][0]
 
    if end_x > start_x + 10:
        return "front"
    elif end_x < start_x - 10:
        return "rear"
    else:
        return "parked"
 
# ================================
# 3. READ IMAGE URLs FROM TXT FILE
# ================================
 
print("🔍 Reading image URLs from urls.txt...")
try:
    with open(URLS_FILE, 'r') as f:
        image_urls = [line.strip() for line in f if line.strip()]
except Exception as e:
    print(f"❌ Failed to read {URLS_FILE}: {e}")
    image_urls = []
 
if not image_urls:
    print("❌ No URLs found in urls.txt")
    exit(1)
 
print(f"✅ Loaded {len(image_urls)} image URLs")
 
# ================================
# 4. SETUP EXCEL OUTPUT
# ================================
 
wb = Workbook()
ws = wb.active
ws.title = "Results"
ws.append(["Image", "Orientation Result"])
 
# Style header
for col in range(1, 3):
    ws.cell(1, col).font = Font(bold=True)
    ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
 
thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)
 
# ================================
# 5. PROCESS EACH IMAGE URL
# ================================
 
for idx, img_url in enumerate(image_urls, start=2):
    print(f"\n--- Processing Image {idx-1}/{len(image_urls)} ---")
    print(f"URL: {img_url}")
 
    # Download image
    try:
        img_data = requests.get(img_url, timeout=10).content
        nparr = np.frombuffer(img_data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            raise ValueError("Failed to decode image")
    except Exception as e:
        error_msg = f"Error downloading image: {str(e)}"
        ws.cell(row=idx, column=2, value=error_msg)
        print(f"❌ {error_msg}")
        continue
 
    # Process frame
    try:
        results = car_model(frame, classes=[2], conf=CONFIDENCE_THRESH)
        detections = results[0].boxes
 
        # Reset tracks for this image (simulate per-image tracking)
        car_tracks.clear()
        next_id = 0
 
        all_results = []
        for i, box in enumerate(detections):
            xyxy = box.xyxy[0].cpu().numpy()
            xmin, ymin, xmax, ymax = map(int, xyxy)
 
            car_crop = frame[ymin:ymax, xmin:xmax]
            if car_crop.size == 0:
                continue
 
            track_id = assign_track_id([xmin, ymin, xmax, ymax])
 
            m1 = get_orientation_dl(car_crop)
            m2 = get_orientation_features(car_crop)
            m3 = get_orientation_motion(track_id)
 
            result = {
                "track_id": int(track_id),
                "method1_dl": m1,
                "method2_features": m2,
                "method3_motion": m3,
                "bbox": [int(xmin), int(ymin), int(xmax), int(ymax)]
            }
            all_results.append(result)
 
            print(f"  Car {i+1} (Track {track_id})")
            print(f"    Method 1 (DL Classifier): {m1}")
            print(f"    Method 2 (Lights/Features): {m2}")
            print(f"    Method 3 (Motion Direction): {m3}")
 
        # Save results as JSON string
        json_text = json.dumps(all_results, indent=2, ensure_ascii=False)
 
        # Embed image into Excel
        pil_img = PILImage.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        pil_img.thumbnail((250, 250))
 
        img_bytes = io.BytesIO()
        pil_img.save(img_bytes, format="PNG")
        img_bytes.seek(0)
 
        xl_img = XLImage(img_bytes)
        ws.row_dimensions[idx].height = 250
        ws.add_image(xl_img, f"A{idx}")
 
        # Write JSON result
        cell = ws.cell(row=idx, column=2, value=json_text)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        cell.border = thin_border
 
        print(f"✅ Success: Processed {img_url}")
 
    except Exception as e:
        error_msg = f"Error processing image: {str(e)}"
        ws.cell(row=idx, column=2, value=error_msg)
        print(f"❌ {error_msg}")
 
# ================================
# 6. SAVE EXCEL FILE
# ================================
 
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 100
OUTPUT_FILE = "model_results_newtrained-file.xlsx"
wb.save(OUTPUT_FILE)
print(f"✅ Results saved to: {OUTPUT_FILE}")